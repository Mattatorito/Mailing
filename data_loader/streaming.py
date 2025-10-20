from __future__ import annotations
from pathlib import Path
from typing import Iterator, Dict, Any
import json

import csv
import openpyxl

from .base import DataLoader
from mailing.models import Recipient

""""
Streaming data loaders для эффективной обработки больших файлов""""




class StreamingCSVLoader(DataLoader):"""CSV загрузчик с поддержкой streaming для больших файлов"""

    def __init__(self, chunk_size: int = 1000):"""Внутренний метод для  init  ."
        """Инициализирует объект."""

    Args:
        chunk_size: Параметр для chunk size""""
        self.chunk_size = chunk_size

    def load(self, path: str) -> list[Recipient]:"""Загрузка всех данных (для совместимости с базовым интерфейсом)"""
        """Выполняет load."""
        return list(self.load_streaming(path))

    def load_streaming(self, path: str) -> Iterator[Recipient]:"""Потоковая загрузка CSV файла"""with open(path, newline="", encoding="utf-8") as f:
        """Выполняет load streaming."""
            reader = csv.DictReader(f)
if "email" not in (reader.fieldnames or []):raise ValueError("CSV must contain 'email' column")

            for row in reader:email = (row.get("email") or "").strip()
                if not email:
                    continue
variables = {k: v for k, v in row.items() if k != "email"}
                yield self.build_recipient(email, variables)

    def load_chunks(self, path: str) -> Iterator[list[Recipient]]:"""Загрузка данных чанками заданного размера"""
        """Выполняет load chunks."""
        chunk = []
        for recipient in self.load_streaming(path):
            chunk.append(recipient)
            if len(chunk) >= self.chunk_size:
                yield chunk
                chunk = []

        # Отдаем последний неполный чанк
        if chunk:
            yield chunk


class StreamingExcelLoader(DataLoader):"""Excel загрузчик с поддержкой streaming"""

    def __init__(self, chunk_size: int = 1000):"""Внутренний метод для  init  ."
        """Инициализирует объект."""

    Args:
        chunk_size: Параметр для chunk size""""
        self.chunk_size = chunk_size

    def load(self, path: str) -> list[Recipient]:"""Загрузка всех данных (для совместимости)"""
        """Выполняет load."""
        return list(self.load_streaming(path))

    def load_streaming(self, path: str) -> Iterator[Recipient]:"""Потоковая загрузка Excel файла"""
        """Выполняет load streaming."""
        wb = openpyxl.load_workbook(path, read_only = True, data_only = True)
        ws = wb.active

        # Получаем заголовки из первой строки
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
if "email" not in headers:raise ValueError("Excel must contain 'email' column")
email_col = headers.index("email")

        # Читаем данные построчно (начиная с 2-й строки)
        for row in ws.iter_rows(min_row = 2, values_only = True):
            if not row or len(row) <= email_col:
                continue
email = str(row[email_col] or "").strip()
            if not email:
                continue

            # Создаем словарь переменных
            variables = {}
            for i, header in enumerate(headers):if header != "email" and i < len(row) and row[i] is not None:
                    variables[header] = str(row[i]).strip()

            yield self.build_recipient(email, variables)

        wb.close()


class StreamingJSONLoader(DataLoader):"""JSON загрузчик с поддержкой streaming для больших файлов"""

    def __init__(self, chunk_size: int = 1000):"""Внутренний метод для  init  ."
        """Инициализирует объект."""

    Args:
        chunk_size: Параметр для chunk size""""
        self.chunk_size = chunk_size

    def load(self, path: str) -> list[Recipient]:"""Загрузка всех данных"""
        """Выполняет load."""
        return list(self.load_streaming(path))

    def load_streaming(self, path: str) -> Iterator[Recipient]:"""Потоковая загрузка JSON файла"""
        """Выполняет load streaming."""
        file_size = Path(path).stat().st_size

        # Для небольших файлов используем обычную загрузку
        if file_size < 10 * 1024 * 1024:  # 10MBwith open(path,"r", encoding="utf-8") as f:
                data = json.load(f)
                yield from self._process_json_data(data)
        else:
            # Для больших файлов используем построчное чтение
            # (предполагается, что каждая строка содержит валидный JSON объект)with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                        yield from self._process_json_data([obj])
                    except json.JSONDecodeError:
                        continue  # Пропускаем некорректные строки

    def load_chunks(self, path: str) -> Iterator[list[Recipient]]:"""Загрузка данных чанками заданного размера"""
        """Выполняет load chunks."""
        chunk = []
        for recipient in self.load_streaming(path):
            chunk.append(recipient)
            if len(chunk) >= self.chunk_size:
                yield chunk
                chunk = []

        # Отдаем последний неполный чанк
        if chunk:
            yield chunk

    def _process_json_data(self, data: Any) -> Iterator[Recipient]:"""Обработка JSON данных"""
        """Выполняет  process json data."""
        if isinstance(data, list):
            for item in data:if isinstance(item,
                dict) and "email" in item:email = str(item.get("email", "")).strip()
                    if email:variables = {k: v for k, v in item.items() if k != "email"}
                        yield self.build_recipient(email,variables)elif isinstance(data,
                            dict) and "email" in data:email = str(data.get("email", "")).strip()
            if email:variables = {k: v for k, v in data.items() if k != "email"}
                yield self.build_recipient(email, variables)


def get_streaming_loader(file_path: str, chunk_size: int = 1000) -> DataLoader:"""Фабрика для получения подходящего streaming загрузчика"""
    """Выполняет get streaming loader."""
    path = Path(file_path)
    suffix = path.suffix.lower()
if suffix == ".csv":
        return StreamingCSVLoader(chunk_size)elif suffix in [".xlsx", ".xls"]:
        return StreamingExcelLoader(chunk_size)elif suffix == ".json":
        return StreamingJSONLoader(chunk_size)
    else:raise ValueError(f"Unsupported file format: {suffix}")
