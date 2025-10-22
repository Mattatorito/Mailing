from __future__ import annotations
from typing import List
from openpyxl import load_workbook
from .base import DataLoader
from mailing.models import Recipient

class ExcelLoader(DataLoader):
    def load(self, path: str) -> List[Recipient]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() for h in rows[0]]
        if 'email' not in headers:
            raise ValueError("XLSX must contain 'email' column")
        email_idx = headers.index('email')
        recipients: List[Recipient] = []
        for row in rows[1:]:
            if row is None:
                continue
            email_val = row[email_idx]
            if not email_val:
                continue
            email = str(email_val).strip()
            variables = {headers[i]: row[i] for i in range(len(headers)) if i != email_idx}
            recipients.append(self.build_recipient(email, variables))
        return recipients
